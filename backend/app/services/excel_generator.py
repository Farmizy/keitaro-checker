"""FB Ads Manager Bulk Upload Excel generator."""

import zoneinfo
from dataclasses import dataclass, field
from datetime import datetime, timedelta

from openpyxl import Workbook

MOSCOW_TZ = zoneinfo.ZoneInfo("Europe/Moscow")

GEO_TO_LANGUAGE = {
    "PL": "Polish",
    "BG": "Bulgarian",
    "RO": "Romanian",
    "LT": "Lithuanian",
    "HU": "Hungarian",
    "CZ": "Czech",
    "HR": "Croatian",
    "SK": "Slovak",
    "SI": "Slovenian",
    "RS": "Serbian",
    "GR": "Greek",
}

BASE_LANGUAGES = ["Albanian", "Chinese (Simplified)", "Georgian"]

ATTRIBUTION_SPEC = (
    '[{"event_type":"CLICK_THROUGH","window_days":1},'
    '{"event_type":"VIEW_THROUGH","window_days":1},'
    '{"event_type":"ENGAGED_VIDEO_VIEW","window_days":1}]'
)

# Ad set name suffixes for copies
ADSET_SUFFIXES = ["", " - Copy", " - Copy 2", " - Copy 3", " - Copy 4"]

# FB Ads Manager Bulk Upload columns (order matches real FB export)
FB_COLUMNS = [
    # Campaign level
    "Campaign Name",
    "Campaign Status",
    "Campaign Objective",
    "Buying Type",
    "Campaign Daily Budget",
    "Campaign Bid Strategy",
    "Campaign Start Time",
    "Campaign Page ID",
    "New Objective",
    # Ad Set level
    "Ad Set Run Status",
    "Ad Set Name",
    "Ad Set Time Start",
    "Destination Type",
    "Link Object ID",
    "Optimized Conversion Tracking Pixels",
    "Optimized Event",
    "Link",
    "Countries",
    "Location Types",
    "Age Min",
    "Age Max",
    "Custom Audiences",
    "Excluded Custom Audiences",
    "Advantage Audience",
    "Individual Setting",
    "Age Range",
    "Targeting Optimization",
    "Beneficiary",
    "Payer",
    "Brand Safety Inventory Filtering Levels",
    "Optimization Goal",
    "Attribution Spec",
    "Billing Event",
    "Regional Regulated Categories",
    # Ad level
    "Ad Status",
    "Ad Name",
    "Dynamic Creative Ad Format",
    "Creative Type",
    "URL Tags",
    "Instagram Account ID (New)",
    "Call to Action",
    "Default Language",
    "Additional Language 1",
    "Additional Language 2",
    "Additional Language 3",
    "Additional Language 4",
]


@dataclass
class CampaignSpec:
    campaign_name: str
    num_adsets: int
    geo: str
    page_id: str
    pixel_id: str
    instagram_id: str
    daily_budget: float
    landing_url: str
    custom_audiences: str
    url_tags: str
    beneficiary: str = ""
    age_min: int = 50
    age_max: int = 65
    default_language: str = "Arabic"
    additional_languages: list[str] = field(default_factory=list)


def generate_fb_excel(specs: list[CampaignSpec]) -> Workbook:
    """Generate FB Ads Manager Bulk Upload Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Upload"

    for col, header in enumerate(FB_COLUMNS, 1):
        ws.cell(row=1, column=col, value=header)

    row_idx = 2
    for spec in specs:
        # Auto-fill languages from geo if not set
        languages = spec.additional_languages
        if not languages:
            geo_lang = GEO_TO_LANGUAGE.get(spec.geo, "")
            languages = BASE_LANGUAGES + ([geo_lang] if geo_lang else [])

        for adset_num in range(spec.num_adsets):
            suffix = (
                ADSET_SUFFIXES[adset_num]
                if adset_num < len(ADSET_SUFFIXES)
                else f" - Copy {adset_num}"
            )
            row_data = _build_row(spec, suffix, adset_num + 1, languages)
            for col, header in enumerate(FB_COLUMNS, 1):
                ws.cell(row=row_idx, column=col, value=row_data.get(header, ""))
            row_idx += 1

    return wb


def _build_row(
    spec: CampaignSpec,
    adset_suffix: str,
    ad_num: int,
    languages: list[str],
) -> dict:
    # Tomorrow at 04:00 AM Moscow time
    tomorrow_4am = (
        datetime.now(MOSCOW_TZ).replace(hour=4, minute=0, second=0, microsecond=0)
        + timedelta(days=1)
    )
    start_time = tomorrow_4am.strftime("%m/%d/%Y %I:%M %p")

    row = {
        # Campaign level
        "Campaign Name": spec.campaign_name,
        "Campaign Status": "PAUSED",
        "Campaign Objective": "Outcome Leads",
        "Buying Type": "AUCTION",
        "Campaign Daily Budget": spec.daily_budget,
        "Campaign Bid Strategy": "Highest volume or value",
        "Campaign Start Time": start_time,
        "Campaign Page ID": f"o:{spec.page_id}",
        "New Objective": "Yes",
        # Ad Set level
        "Ad Set Run Status": "ACTIVE",
        "Ad Set Name": f"New Leads Ad Set{adset_suffix}",
        "Ad Set Time Start": start_time,
        "Destination Type": "UNDEFINED",
        "Link Object ID": f"o:{spec.page_id}",
        "Optimized Conversion Tracking Pixels": f"tp:{spec.pixel_id}",
        "Optimized Event": "LEAD",
        "Link": spec.landing_url,
        "Countries": spec.geo,
        "Location Types": "home, recent",
        "Age Min": spec.age_min,
        "Age Max": spec.age_max,
        "Custom Audiences": spec.custom_audiences,
        "Excluded Custom Audiences": spec.custom_audiences,
        "Advantage Audience": 1,
        "Individual Setting": "age: On, gender: On",
        "Age Range": f"{spec.age_min}, {spec.age_max}",
        "Targeting Optimization": "expansion_all",
        "Beneficiary": spec.beneficiary,
        "Payer": spec.beneficiary,
        "Brand Safety Inventory Filtering Levels": "FACEBOOK_RELAXED, AN_RELAXED",
        "Optimization Goal": "OFFSITE_CONVERSIONS",
        "Attribution Spec": ATTRIBUTION_SPEC,
        "Billing Event": "IMPRESSIONS",
        "Regional Regulated Categories": "VOLUNTARY_VERIFICATION",
        # Ad level
        "Ad Status": "ACTIVE",
        "Ad Name": str(ad_num),
        "Dynamic Creative Ad Format": "Automatic Format",
        "Creative Type": "Link Page Post Ad",
        "URL Tags": spec.url_tags,
        "Instagram Account ID (New)": (
            f"x:{spec.instagram_id}" if spec.instagram_id else ""
        ),
        "Call to Action": "LEARN_MORE",
        "Default Language": spec.default_language,
    }

    for i, lang in enumerate(languages):
        row[f"Additional Language {i + 1}"] = lang

    return row
